#!/usr/bin/env python3
"""Pharma sales analytics from Excel for business reporting and dashboarding."""
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

EXCEL_FILE = Path("Pharmaceutical_Drug_Sales.xlsx")
RAW_SHEET = "Raw_Data "
HELPER_SHEET = "Helper_Table"

CSV_OUTPUTS = {
    "cleaned_data": Path("cleaned_data.csv"),
    "region_summary": Path("region_summary.csv"),
    "monthly_summary": Path("monthly_summary.csv"),
    "rep_summary": Path("sales_rep_summary.csv"),
    "product_summary": Path("product_summary.csv"),
}


def parse_sheet(sheet_name: str) -> Tuple[List[str], List[Tuple]]:
    workbook = load_workbook(EXCEL_FILE, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet '{sheet_name}' is empty")
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    return header, [tuple(row) for row in rows[1:] if any(cell is not None for cell in row)]


def parse_date(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"]:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def month_label(dt: Optional[datetime]) -> str:
    return dt.strftime("%b") if dt else "Unknown"


def quarter_label(dt: Optional[datetime]) -> str:
    if not dt:
        return "Unknown"
    quarter = (dt.month - 1) // 3 + 1
    return f"Q{quarter}"


def revenue_category(revenue: float) -> str:
    if revenue >= 40000:
        return "High"
    if revenue >= 20000:
        return "Medium"
    return "Low"


def load_helper_map() -> Dict[str, float]:
    header, rows = parse_sheet(HELPER_SHEET)
    region_index = header.index("Region")
    multiplier_index = header.index("Target Multiplier")
    helper_map = {}
    for row in rows:
        region = row[region_index]
        multiplier = row[multiplier_index]
        if region and multiplier:
            helper_map[str(region).strip()] = float(multiplier)
    return helper_map


def build_cleaned_rows(raw_header: List[str], raw_rows: List[Tuple], helper_map: Dict[str, float]) -> Tuple[List[str], List[Tuple]]:
    header = raw_header + [
        "Month",
        "Quarter",
        "Revenue Category",
        "Target Multiplier",
        "Target Revenue",
        "Achievement Percentage",
        "Performance Status",
    ]
    region_index = raw_header.index("Region")
    revenue_index = raw_header.index("Revenue")
    sale_date_index = raw_header.index("Sale Date")

    cleaned_rows = []
    for row in raw_rows:
        revenue = float(row[revenue_index] or 0)
        sale_date = parse_date(row[sale_date_index])
        region = str(row[region_index]).strip() if row[region_index] else "Unknown"
        multiplier = helper_map.get(region, 1.0)
        target_revenue = revenue * multiplier
        achievement_pct = revenue / target_revenue if target_revenue else 0.0
        performance_status = "Achieved" if achievement_pct >= 1.0 else "Below Target"

        cleaned_rows.append(
            row
            + (
                month_label(sale_date),
                quarter_label(sale_date),
                revenue_category(revenue),
                multiplier,
                round(target_revenue, 2),
                round(achievement_pct, 4),
                performance_status,
            )
        )
    return header, cleaned_rows


def write_csv(path: Path, header: List[str], rows: List[Tuple]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fp:
        fp.write(",".join(f'"{col}"' for col in header) + "\n")
        for row in rows:
            row_values = ["" if v is None else str(v) for v in row]
            escaped = [value.replace('"', '""') for value in row_values]
            fp.write(",".join(f'"{value}"' for value in escaped) + "\n")


def numeric(value):
    try:
        return float(value)
    except Exception:
        return 0.0


def summarize(cleaned_header: List[str], cleaned_rows: List[Tuple]) -> None:
    region_idx = cleaned_header.index("Region")
    revenue_idx = cleaned_header.index("Revenue")
    units_idx = cleaned_header.index("Units Sold")
    achievement_idx = cleaned_header.index("Achievement Percentage")
    drug_idx = cleaned_header.index("Drug Name")
    rep_idx = cleaned_header.index("Sales Representative")
    month_idx = cleaned_header.index("Month")

    summary = {
        "total_revenue": 0.0,
        "total_units": 0,
        "sum_achievement": 0.0,
        "row_count": 0,
    }
    revenue_by_region = defaultdict(float)
    revenue_by_month = defaultdict(float)
    revenue_by_product = defaultdict(float)
    revenue_by_rep = defaultdict(float)

    for row in cleaned_rows:
        revenue = numeric(row[revenue_idx])
        units = numeric(row[units_idx])
        achievement = numeric(row[achievement_idx])
        summary["total_revenue"] += revenue
        summary["total_units"] += units
        summary["sum_achievement"] += achievement
        summary["row_count"] += 1
        revenue_by_region[row[region_idx]] += revenue
        revenue_by_month[row[month_idx]] += revenue
        revenue_by_product[row[drug_idx]] += revenue
        revenue_by_rep[row[rep_idx]] += revenue

    average_achievement = summary["sum_achievement"] / summary["row_count"] if summary["row_count"] else 0.0
    top_region = max(revenue_by_region.items(), key=lambda t: t[1]) if revenue_by_region else ("", 0.0)
    top_product = max(revenue_by_product.items(), key=lambda t: t[1]) if revenue_by_product else ("", 0.0)

    print("\n=== Commercial Analytics Summary ===")
    print(f"Total Revenue: ${summary['total_revenue']:,.2f}")
    print(f"Total Units Sold: {int(summary['total_units']):,}")
    print(f"Average Achievement %: {average_achievement:.2%}")
    print(f"Top Performing Region: {top_region[0]} (${top_region[1]:,.2f})")
    print(f"Top Selling Drug: {top_product[0]} (${top_product[1]:,.2f})")

    print("\nRevenue by Region")
    for region, revenue in sorted(revenue_by_region.items(), key=lambda x: x[1], reverse=True):
        print(f"  {region}: ${revenue:,.2f}")

    print("\nRevenue by Month")
    for month, revenue in sorted(revenue_by_month.items(), key=lambda x: datetime.strptime(x[0], "%b").month if x[0] != "Unknown" else 0):
        print(f"  {month}: ${revenue:,.2f}")

    print("\nTop 10 Sales Representatives")
    for rep, revenue in sorted(revenue_by_rep.items(), key=lambda x: x[1], reverse=True)[:10]:
        print(f"  {rep}: ${revenue:,.2f}")

    write_summary_tables(cleaned_header, cleaned_rows)


def write_summary_tables(cleaned_header: List[str], cleaned_rows: List[Tuple]) -> None:
    region_idx = cleaned_header.index("Region")
    revenue_idx = cleaned_header.index("Revenue")
    units_idx = cleaned_header.index("Units Sold")
    achievement_idx = cleaned_header.index("Achievement Percentage")
    drug_idx = cleaned_header.index("Drug Name")
    rep_idx = cleaned_header.index("Sales Representative")
    month_idx = cleaned_header.index("Month")

    region_data = defaultdict(lambda: [0.0, 0.0, 0])
    month_data = defaultdict(lambda: [0.0, 0.0, 0])
    rep_data = defaultdict(lambda: [0.0, 0, 0.0])
    product_data = defaultdict(lambda: [0.0, 0])

    for row in cleaned_rows:
        revenue = numeric(row[revenue_idx])
        units = numeric(row[units_idx])
        achievement = numeric(row[achievement_idx])
        region = row[region_idx]
        month = row[month_idx]
        rep = row[rep_idx]
        product = row[drug_idx]

        region_data[region][0] += revenue
        region_data[region][1] += units
        region_data[region][2] += 1

        month_data[month][0] += revenue
        month_data[month][1] += units
        month_data[month][2] += 1

        rep_data[rep][0] += revenue
        rep_data[rep][1] += 1
        rep_data[rep][2] += achievement

        product_data[product][0] += revenue
        product_data[product][1] += 1

    write_csv(
        CSV_OUTPUTS["region_summary"],
        ["Region", "Total Revenue", "Total Units Sold", "Transactions"],
        [
            (region, round(values[0], 2), int(values[1]), int(values[2]))
            for region, values in sorted(region_data.items(), key=lambda item: item[1][0], reverse=True)
        ],
    )
    write_csv(
        CSV_OUTPUTS["monthly_summary"],
        ["Month", "Total Revenue", "Total Units Sold", "Transactions"],
        [
            (month, round(values[0], 2), int(values[1]), int(values[2]))
            for month, values in sorted(month_data.items(), key=lambda item: datetime.strptime(item[0], "%b").month if item[0] != "Unknown" else 0)
        ],
    )
    write_csv(
        CSV_OUTPUTS["rep_summary"],
        ["Sales Representative", "Total Revenue", "Transactions", "Average Achievement %"],
        [
            (rep, round(values[0], 2), int(values[1]), round(values[2] / values[1], 4) if values[1] else 0.0)
            for rep, values in sorted(rep_data.items(), key=lambda item: item[1][0], reverse=True)
        ],
    )
    write_csv(
        CSV_OUTPUTS["product_summary"],
        ["Drug Name", "Total Revenue", "Transactions"],
        [
            (product, round(values[0], 2), int(values[1]))
            for product, values in sorted(product_data.items(), key=lambda item: item[1][0], reverse=True)
        ],
    )


def main() -> None:
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Could not find {EXCEL_FILE}")

    raw_header, raw_rows = parse_sheet(RAW_SHEET)
    helper_map = load_helper_map()
    cleaned_header, cleaned_rows = build_cleaned_rows(raw_header, raw_rows, helper_map)

    write_csv(CSV_OUTPUTS["cleaned_data"], cleaned_header, cleaned_rows)
    print(f"Cleaned dataset saved to {CSV_OUTPUTS['cleaned_data']}")
    summarize(cleaned_header, cleaned_rows)


if __name__ == "__main__":
    main()
