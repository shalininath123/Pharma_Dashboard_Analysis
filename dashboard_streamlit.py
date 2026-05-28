#!/usr/bin/env python3
"""Interactive Pharma Sales Dashboard using Streamlit and Plotly."""
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook

st.set_page_config(page_title="Pharma Sales Dashboard", layout="wide", initial_sidebar_state="expanded")

EXCEL_FILE = Path("Pharmaceutical_Drug_Sales.xlsx")
RAW_SHEET = "Raw_Data "
HELPER_SHEET = "Helper_Table"


@st.cache_data
def load_data():
    """Load and process data from Excel."""
    workbook = load_workbook(EXCEL_FILE, data_only=True)
    raw_sheet = workbook[RAW_SHEET]
    helper_sheet = workbook[HELPER_SHEET]

    rows = list(raw_sheet.iter_rows(values_only=True))
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    raw_data = [tuple(row) for row in rows[1:] if any(cell is not None for cell in row)]

    helper_rows = list(helper_sheet.iter_rows(values_only=True))
    helper_header = [str(cell).strip() if cell is not None else "" for cell in helper_rows[0]]
    helper_data = [tuple(row) for row in helper_rows[1:] if any(cell is not None for cell in row)]

    region_idx = helper_header.index("Region") if "Region" in helper_header else 0
    multiplier_idx = helper_header.index("Target Multiplier") if "Target Multiplier" in helper_header else 1
    helper_map = {str(row[region_idx]).strip(): float(row[multiplier_idx]) for row in helper_data if row[region_idx] and row[multiplier_idx]}

    df_list = []
    for row in raw_data:
        row_dict = {header[i]: row[i] for i in range(len(header))}
        row_dict["Revenue"] = float(row_dict.get("Revenue") or 0)
        row_dict["Units Sold"] = int(float(row_dict.get("Units Sold") or 0))
        
        sale_date = row_dict.get("Sale Date")
        if isinstance(sale_date, datetime):
            row_dict["Sale Date"] = sale_date
            row_dict["Month"] = sale_date.strftime("%b")
            quarter = (sale_date.month - 1) // 3 + 1
            row_dict["Quarter"] = f"Q{quarter}"
        else:
            row_dict["Month"] = "Unknown"
            row_dict["Quarter"] = "Unknown"

        region = str(row_dict.get("Region") or "Unknown").strip()
        multiplier = helper_map.get(region, 1.0)
        target_revenue = row_dict["Revenue"] * multiplier
        row_dict["Target Multiplier"] = multiplier
        row_dict["Target Revenue"] = target_revenue
        row_dict["Achievement %"] = (row_dict["Revenue"] / target_revenue * 100) if target_revenue else 0
        row_dict["Performance Status"] = "Achieved" if row_dict["Achievement %"] >= 100 else "Below Target"

        df_list.append(row_dict)

    return pd.DataFrame(df_list)


def main():
    st.title("💊 Pharmaceutical Sales & Commercial Analytics Dashboard")
    st.markdown("---")

    df = load_data()

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("Total Revenue", f"${df['Revenue'].sum() / 1_000_000:.2f}M", delta=f"{df['Achievement %'].mean():.1f}%")
    with col2:
        st.metric("Total Units Sold", f"{df['Units Sold'].sum():,}")
    with col3:
        st.metric("Avg Achievement %", f"{df['Achievement %'].mean():.1f}%")
    with col4:
        top_region = df.groupby("Region")["Revenue"].sum().idxmax()
        st.metric("Top Region", top_region)
    with col5:
        top_drug = df.groupby("Drug Name")["Revenue"].sum().idxmax()
        st.metric("Top Drug", top_drug)

    st.markdown("---")

    st.sidebar.header("🔍 Filters")
    regions = ["All"] + sorted(df["Region"].unique().tolist())
    selected_region = st.sidebar.selectbox("Select Region", regions)
    
    drugs = ["All"] + sorted(df["Drug Name"].unique().tolist())
    selected_drug = st.sidebar.selectbox("Select Drug", drugs)

    months = ["All"] + sorted(df["Month"].unique().tolist(), key=lambda x: ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"].index(x) if x != "Unknown" else 999)
    selected_month = st.sidebar.selectbox("Select Month", months)

    status_options = ["All"] + sorted(df["Performance Status"].unique().tolist())
    selected_status = st.sidebar.selectbox("Select Performance Status", status_options)

    df_filtered = df.copy()
    if selected_region != "All":
        df_filtered = df_filtered[df_filtered["Region"] == selected_region]
    if selected_drug != "All":
        df_filtered = df_filtered[df_filtered["Drug Name"] == selected_drug]
    if selected_month != "All":
        df_filtered = df_filtered[df_filtered["Month"] == selected_month]
    if selected_status != "All":
        df_filtered = df_filtered[df_filtered["Performance Status"] == selected_status]

    st.sidebar.markdown("---")
    st.sidebar.info(f"📊 Filtered Data: {len(df_filtered)} records out of {len(df)}")

    tab1, tab2, tab3, tab4 = st.tabs(["Overview", "Regional Analysis", "Product Analysis", "Performance"])

    with tab1:
        col1, col2 = st.columns(2)
        with col1:
            fig_revenue_month = px.line(
                df_filtered.groupby("Month")["Revenue"].sum().reset_index(),
                x="Month",
                y="Revenue",
                title="Monthly Revenue Trend",
                markers=True,
                labels={"Revenue": "Total Revenue ($)"}
            )
            fig_revenue_month.update_layout(height=400)
            st.plotly_chart(fig_revenue_month, use_container_width=True)

        with col2:
            fig_units_month = px.bar(
                df_filtered.groupby("Month")["Units Sold"].sum().reset_index(),
                x="Month",
                y="Units Sold",
                title="Monthly Units Sold",
                labels={"Units Sold": "Total Units"}
            )
            fig_units_month.update_layout(height=400)
            st.plotly_chart(fig_units_month, use_container_width=True)

        col3, col4 = st.columns(2)
        with col3:
            status_dist = df_filtered["Performance Status"].value_counts().reset_index()
            status_dist.columns = ["Performance Status", "Count"]
            fig_status = px.pie(
                status_dist,
                names="Performance Status",
                values="Count",
                title="Performance Status Distribution",
                color_discrete_map={"Achieved": "#34a853", "Below Target": "#ea4335"}
            )
            fig_status.update_layout(height=400)
            st.plotly_chart(fig_status, use_container_width=True)

        with col4:
            customer_dist = df_filtered["Customer Type"].value_counts().reset_index()
            customer_dist.columns = ["Customer Type", "Count"]
            fig_customer = px.bar(
                customer_dist,
                x="Customer Type",
                y="Count",
                title="Customer Type Distribution",
                labels={"Count": "Transaction Count"}
            )
            fig_customer.update_layout(height=400)
            st.plotly_chart(fig_customer, use_container_width=True)

    with tab2:
        col1, col2 = st.columns(2)
        with col1:
            region_revenue = df_filtered.groupby("Region")["Revenue"].sum().reset_index().sort_values("Revenue", ascending=True)
            fig_region = px.bar(
                region_revenue,
                x="Revenue",
                y="Region",
                orientation="h",
                title="Revenue by Region",
                labels={"Revenue": "Total Revenue ($)"}
            )
            fig_region.update_layout(height=450)
            st.plotly_chart(fig_region, use_container_width=True)

        with col2:
            region_units = df_filtered.groupby("Region")["Units Sold"].sum().reset_index().sort_values("Units Sold", ascending=True)
            fig_region_units = px.bar(
                region_units,
                x="Units Sold",
                y="Region",
                orientation="h",
                title="Units Sold by Region",
                labels={"Units Sold": "Total Units"}
            )
            fig_region_units.update_layout(height=450)
            st.plotly_chart(fig_region_units, use_container_width=True)

        region_summary = df_filtered.groupby("Region").agg({
            "Revenue": "sum",
            "Units Sold": "sum",
            "Achievement %": "mean"
        }).round(2).reset_index()
        region_summary.columns = ["Region", "Total Revenue", "Total Units", "Avg Achievement %"]
        st.subheader("📊 Regional Summary")
        st.dataframe(region_summary, use_container_width=True)

    with tab3:
        col1, col2 = st.columns(2)
        with col1:
            drug_revenue = df_filtered.groupby("Drug Name")["Revenue"].sum().reset_index().sort_values("Revenue", ascending=True).tail(15)
            fig_drug = px.bar(
                drug_revenue,
                x="Revenue",
                y="Drug Name",
                orientation="h",
                title="Top 15 Drugs by Revenue",
                labels={"Revenue": "Total Revenue ($)"}
            )
            fig_drug.update_layout(height=500)
            st.plotly_chart(fig_drug, use_container_width=True)

        with col2:
            drug_units = df_filtered.groupby("Drug Name")["Units Sold"].sum().reset_index().sort_values("Units Sold", ascending=True).tail(15)
            fig_drug_units = px.bar(
                drug_units,
                x="Units Sold",
                y="Drug Name",
                orientation="h",
                title="Top 15 Drugs by Units Sold",
                labels={"Units Sold": "Total Units"}
            )
            fig_drug_units.update_layout(height=500)
            st.plotly_chart(fig_drug_units, use_container_width=True)

        drug_summary = df_filtered.groupby("Drug Name").agg({
            "Revenue": "sum",
            "Units Sold": "sum",
            "Achievement %": "mean"
        }).round(2).reset_index().sort_values("Revenue", ascending=False)
        drug_summary.columns = ["Drug Name", "Total Revenue", "Total Units", "Avg Achievement %"]
        st.subheader("📊 Product Summary")
        st.dataframe(drug_summary, use_container_width=True)

    with tab4:
        col1, col2 = st.columns(2)
        with col1:
            achievement_by_region = df_filtered.groupby("Region")["Achievement %"].mean().reset_index().sort_values("Achievement %")
            fig_achievement = px.bar(
                achievement_by_region,
                x="Region",
                y="Achievement %",
                title="Avg Achievement % by Region",
                labels={"Achievement %": "Achievement Percentage (%)"}
            )
            fig_achievement.update_layout(height=400)
            st.plotly_chart(fig_achievement, use_container_width=True)

        with col2:
            sales_rep_performance = df_filtered.groupby("Sales Representative").agg({
                "Revenue": "sum",
                "Achievement %": "mean"
            }).reset_index().sort_values("Revenue", ascending=False).head(10)
            fig_rep = px.scatter(
                sales_rep_performance,
                x="Achievement %",
                y="Revenue",
                size="Revenue",
                hover_name="Sales Representative",
                title="Top 10 Sales Reps: Revenue vs Achievement %",
                labels={"Achievement %": "Avg Achievement %", "Revenue": "Total Revenue ($)"}
            )
            fig_rep.update_layout(height=400)
            st.plotly_chart(fig_rep, use_container_width=True)

        st.subheader("📈 Sales Representative Performance")
        rep_summary = df_filtered.groupby("Sales Representative").agg({
            "Revenue": "sum",
            "Units Sold": "sum",
            "Achievement %": "mean"
        }).round(2).reset_index().sort_values("Revenue", ascending=False)
        rep_summary.columns = ["Sales Representative", "Total Revenue", "Total Units", "Avg Achievement %"]
        st.dataframe(rep_summary, use_container_width=True)

        st.subheader("🎯 Target Achievement Distribution")
        achievement_dist = pd.cut(df_filtered["Achievement %"], bins=[0, 50, 75, 100, 150], labels=["0-50%", "50-75%", "75-100%", "100%+"]).value_counts().reset_index()
        achievement_dist.columns = ["Achievement Range", "Count"]
        fig_achievement_dist = px.bar(
            achievement_dist,
            x="Achievement Range",
            y="Count",
            title="Achievement Distribution",
            color="Achievement Range"
        )
        fig_achievement_dist.update_layout(height=350)
        st.plotly_chart(fig_achievement_dist, use_container_width=True)

    st.markdown("---")
    st.markdown("### 📋 Raw Data Preview")
    st.dataframe(df_filtered.head(20), use_container_width=True)


if __name__ == "__main__":
    main()
